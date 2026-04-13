import logging
import os
import sqlite3
import tempfile
from calendar import monthrange
from datetime import datetime, timedelta, time
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from telegram import InputFile, ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")
DB_NAME = "tracker_bot.db"
LOCAL_TZ = ZoneInfo("Europe/Riga")
CURRENCY = "EUR"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

WORK_HOURS, WORK_LOCATION = range(100, 102)
EXPENSE_GROUP, EXPENSE_AMOUNT, EXPENSE_CATEGORY, EXPENSE_NOTE = range(200, 204)

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["🕒 Ish soati qo'shish", "💸 Xarajat qo'shish"],
        ["📅 Bugun", "📊 Hafta"],
        ["🗓 Oy", "🧾 Oxirgi xarajatlar"],
        ["📤 Haftalik export", "📤 Oylik export"],
        ["❌ Bekor qilish"],
    ],
    resize_keyboard=True,
)

EXPENSE_GROUP_KEYBOARD = ReplyKeyboardMarkup(
    [["🏠 Uy uchun", "👤 O'zim uchun"], ["❌ Bekor qilish"]],
    resize_keyboard=True,
    one_time_keyboard=True,
)


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            chat_id INTEGER PRIMARY KEY,
            full_name TEXT,
            username TEXT,
            created_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS work_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER NOT NULL,
            work_date TEXT NOT NULL,
            hours REAL NOT NULL,
            location TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (chat_id) REFERENCES users(chat_id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER NOT NULL,
            expense_group TEXT NOT NULL,
            amount REAL NOT NULL,
            category TEXT NOT NULL,
            note TEXT,
            spent_at TEXT NOT NULL,
            FOREIGN KEY (chat_id) REFERENCES users(chat_id)
        )
        """
    )

    conn.commit()
    conn.close()


def now_local() -> datetime:
    return datetime.now(LOCAL_TZ)


def parse_iso(value: str) -> datetime:
    return datetime.fromisoformat(value)


def ensure_user(chat_id: int, full_name: str, username: str | None) -> None:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO users (chat_id, full_name, username, created_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(chat_id) DO UPDATE SET
            full_name = excluded.full_name,
            username = excluded.username
        """,
        (chat_id, full_name, username, now_local().isoformat()),
    )
    conn.commit()
    conn.close()


def get_all_user_ids() -> list[int]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT chat_id FROM users ORDER BY chat_id ASC")
    rows = cur.fetchall()
    conn.close()
    return [int(row["chat_id"]) for row in rows]


def add_work_entry(chat_id: int, hours: float, location: str, created_at: datetime) -> None:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO work_entries (chat_id, work_date, hours, location, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            chat_id,
            created_at.date().isoformat(),
            hours,
            location.strip(),
            created_at.isoformat(),
        ),
    )
    conn.commit()
    conn.close()


def fetch_work_entries_between(chat_id: int, start_dt: datetime, end_dt: datetime):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, work_date, hours, location, created_at
        FROM work_entries
        WHERE chat_id = ?
          AND created_at >= ?
          AND created_at < ?
        ORDER BY created_at ASC
        """,
        (chat_id, start_dt.isoformat(), end_dt.isoformat()),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def add_expense(
    chat_id: int,
    expense_group: str,
    amount: float,
    category: str,
    note: str | None,
    spent_at: datetime,
) -> None:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO expenses (chat_id, expense_group, amount, category, note, spent_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (chat_id, expense_group.strip(), amount, category.strip(), (note or "").strip(), spent_at.isoformat()),
    )
    conn.commit()
    conn.close()


def fetch_expenses_between(chat_id: int, start_dt: datetime, end_dt: datetime):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, expense_group, amount, category, note, spent_at
        FROM expenses
        WHERE chat_id = ?
          AND spent_at >= ?
          AND spent_at < ?
        ORDER BY spent_at DESC
        """,
        (chat_id, start_dt.isoformat(), end_dt.isoformat()),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_recent_expenses(chat_id: int, limit: int = 10):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT expense_group, amount, category, note, spent_at
        FROM expenses
        WHERE chat_id = ?
        ORDER BY spent_at DESC
        LIMIT ?
        """,
        (chat_id, limit),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def today_range(base_dt: datetime):
    start_dt = datetime.combine(base_dt.date(), time.min, tzinfo=LOCAL_TZ)
    end_dt = start_dt + timedelta(days=1)
    return start_dt, end_dt


def week_range(base_dt: datetime):
    start_date = base_dt.date() - timedelta(days=base_dt.weekday())
    start_dt = datetime.combine(start_date, time.min, tzinfo=LOCAL_TZ)
    end_dt = start_dt + timedelta(days=7)
    return start_dt, end_dt


def month_range(base_dt: datetime):
    start_dt = datetime(base_dt.year, base_dt.month, 1, tzinfo=LOCAL_TZ)
    if base_dt.month == 12:
        end_dt = datetime(base_dt.year + 1, 1, 1, tzinfo=LOCAL_TZ)
    else:
        end_dt = datetime(base_dt.year, base_dt.month + 1, 1, tzinfo=LOCAL_TZ)
    return start_dt, end_dt


def format_minutes(total_minutes: int) -> str:
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours} soat {minutes} daqiqa"


def format_hours(hours: float) -> str:
    total_minutes = int(round(hours * 60))
    return format_minutes(total_minutes)


def format_dt(dt: datetime) -> str:
    return dt.astimezone(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")


def last_day_of_month(dt: datetime) -> int:
    return monthrange(dt.year, dt.month)[1]


def build_summary_data(chat_id: int, start_dt: datetime, end_dt: datetime) -> dict:
    work_entries = fetch_work_entries_between(chat_id, start_dt, end_dt)
    expenses = fetch_expenses_between(chat_id, start_dt, end_dt)

    total_work_hours = 0.0
    location_totals: dict[str, float] = {}
    daily_breakdown: dict[str, dict] = {}

    iter_day = start_dt.date()
    while iter_day < end_dt.date():
        daily_breakdown[iter_day.isoformat()] = {
            "work_hours": 0.0,
            "uy_total": 0.0,
            "ozim_total": 0.0,
        }
        iter_day += timedelta(days=1)

    for row in work_entries:
        hours = float(row["hours"])
        total_work_hours += hours
        location = row["location"]
        location_totals[location] = location_totals.get(location, 0.0) + hours

        key = row["work_date"]
        if key in daily_breakdown:
            daily_breakdown[key]["work_hours"] += hours

    total_expense = 0.0
    expense_group_totals = {"Uy uchun": 0.0, "O'zim uchun": 0.0}
    category_totals: dict[str, float] = {}

    for row in expenses:
        amount = float(row["amount"])
        total_expense += amount
        expense_group = row["expense_group"]
        expense_group_totals[expense_group] = expense_group_totals.get(expense_group, 0.0) + amount
        category_totals[row["category"]] = category_totals.get(row["category"], 0.0) + amount

        key = parse_iso(row["spent_at"]).date().isoformat()
        if key in daily_breakdown:
            if expense_group == "Uy uchun":
                daily_breakdown[key]["uy_total"] += amount
            else:
                daily_breakdown[key]["ozim_total"] += amount

    top_category = None
    if category_totals:
        top_category = max(category_totals.items(), key=lambda x: x[1])

    return {
        "start_dt": start_dt,
        "end_dt": end_dt,
        "work_entries": work_entries,
        "expenses": expenses,
        "total_work_hours": total_work_hours,
        "location_totals": location_totals,
        "total_expense": total_expense,
        "expense_group_totals": expense_group_totals,
        "category_totals": category_totals,
        "top_category": top_category,
        "daily_breakdown": daily_breakdown,
    }


def format_summary(title: str, data: dict) -> str:
    lines = [title, ""]
    uy_total = data["expense_group_totals"].get("Uy uchun", 0.0)
    ozim_total = data["expense_group_totals"].get("O'zim uchun", 0.0)

    lines.append(f"🕒 Jami ish vaqti: {format_hours(data['total_work_hours'])}")
    lines.append(f"📝 Kiritilgan ish yozuvlari: {len(data['work_entries'])} ta")
    lines.append(f"💸 Jami xarajat: {data['total_expense']:.2f} {CURRENCY}")
    lines.append(f"🏠 Uy uchun xarajat: {uy_total:.2f} {CURRENCY}")
    lines.append(f"👤 O'zim uchun xarajat: {ozim_total:.2f} {CURRENCY}")

    if data["top_category"]:
        lines.append(
            f"🔥 Eng ko'p xarajat kategoriyasi: {data['top_category'][0]} ({data['top_category'][1]:.2f} {CURRENCY})"
        )

    if data["location_totals"]:
        lines.append("")
        lines.append("📍 Qayerda ishlaganim:")
        for location, hours in sorted(data["location_totals"].items(), key=lambda x: x[1], reverse=True):
            lines.append(f"- {location}: {format_hours(hours)}")

    if data["category_totals"]:
        lines.append("")
        lines.append("🧾 Kategoriya bo'yicha xarajatlar:")
        for category, amount in sorted(data["category_totals"].items(), key=lambda x: x[1], reverse=True):
            lines.append(f"- {category}: {amount:.2f} {CURRENCY}")

    if data["daily_breakdown"]:
        lines.append("")
        lines.append("📅 Kunlar bo'yicha breakdown:")
        for date_key, item in data["daily_breakdown"].items():
            if item["work_hours"] == 0 and item["uy_total"] == 0 and item["ozim_total"] == 0:
                continue
            lines.append(
                f"- {date_key}: {format_hours(item['work_hours'])} | Uy: {item['uy_total']:.2f} {CURRENCY} | O'zim: {item['ozim_total']:.2f} {CURRENCY}"
            )

    return "\n".join(lines)


def build_export_basename(period_name: str, start_dt: datetime, chat_id: int) -> str:
    return f"tracker_{period_name}_{chat_id}_{start_dt.strftime('%Y%m%d')}"


def create_excel_report(chat_id: int, title: str, data: dict, period_name: str) -> Path:
    temp_dir = Path(tempfile.gettempdir())
    file_path = temp_dir / f"{build_export_basename(period_name, data['start_dt'], chat_id)}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    ws.append([title])
    ws.append([])
    ws.append(["Jami ish vaqti", format_hours(data["total_work_hours"])])
    ws.append(["Ish yozuvlari soni", len(data["work_entries"])])
    ws.append(["Jami xarajat", data["total_expense"]])
    ws.append(["Uy uchun", data["expense_group_totals"].get("Uy uchun", 0.0)])
    ws.append(["O'zim uchun", data["expense_group_totals"].get("O'zim uchun", 0.0)])
    ws.append(["Valyuta", CURRENCY])
    if data["top_category"]:
        ws.append(["Eng ko'p kategoriya", data["top_category"][0]])
        ws.append(["Eng ko'p kategoriya summasi", data["top_category"][1]])

    ws2 = wb.create_sheet("Kunlar")
    ws2.append(["Sana", "Ish vaqti (soat)", "Uy uchun", "O'zim uchun"])
    for date_key, item in data["daily_breakdown"].items():
        ws2.append([date_key, item["work_hours"], item["uy_total"], item["ozim_total"]])

    ws3 = wb.create_sheet("Ish yozuvlari")
    ws3.append(["Kiritilgan vaqt", "Sana", "Joy", "Ish vaqti (soat)"])
    for row in data["work_entries"]:
        ws3.append([
            format_dt(parse_iso(row["created_at"])),
            row["work_date"],
            row["location"],
            float(row["hours"]),
        ])

    ws4 = wb.create_sheet("Xarajatlar")
    ws4.append(["Vaqt", "Guruh", "Kategoriya", "Izoh", f"Summa ({CURRENCY})"])
    for row in data["expenses"]:
        ws4.append([
            format_dt(parse_iso(row["spent_at"])),
            row["expense_group"],
            row["category"],
            row["note"],
            float(row["amount"]),
        ])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_len = 0
            column_letter = column[0].column_letter
            for cell in column:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 40)

    wb.save(file_path)
    return file_path


def draw_multiline_pdf_text(pdf: canvas.Canvas, lines: list[str], x: int = 40, y: int = 800) -> None:
    text_obj = pdf.beginText(x, y)
    text_obj.setFont("Helvetica", 11)
    for line in lines:
        if text_obj.getY() < 50:
            pdf.drawText(text_obj)
            pdf.showPage()
            text_obj = pdf.beginText(x, 800)
            text_obj.setFont("Helvetica", 11)
        text_obj.textLine(line)
    pdf.drawText(text_obj)


def create_pdf_report(chat_id: int, title: str, data: dict, period_name: str) -> Path:
    temp_dir = Path(tempfile.gettempdir())
    file_path = temp_dir / f"{build_export_basename(period_name, data['start_dt'], chat_id)}.pdf"
    pdf = canvas.Canvas(str(file_path), pagesize=A4)

    uy_total = data["expense_group_totals"].get("Uy uchun", 0.0)
    ozim_total = data["expense_group_totals"].get("O'zim uchun", 0.0)

    lines = [
        title,
        "",
        f"Jami ish vaqti: {format_hours(data['total_work_hours'])}",
        f"Ish yozuvlari soni: {len(data['work_entries'])} ta",
        f"Jami xarajat: {data['total_expense']:.2f} {CURRENCY}",
        f"Uy uchun: {uy_total:.2f} {CURRENCY}",
        f"O'zim uchun: {ozim_total:.2f} {CURRENCY}",
    ]

    if data["top_category"]:
        lines.append(
            f"Eng ko'p kategoriya: {data['top_category'][0]} ({data['top_category'][1]:.2f} {CURRENCY})"
        )

    lines.append("")
    lines.append("Qayerda ishlaganim:")
    if data["location_totals"]:
        for location, hours in sorted(data["location_totals"].items(), key=lambda x: x[1], reverse=True):
            lines.append(f"- {location}: {format_hours(hours)}")
    else:
        lines.append("- Ma'lumot yo'q")

    lines.append("")
    lines.append("Kategoriya bo'yicha xarajatlar:")
    if data["category_totals"]:
        for category, amount in sorted(data["category_totals"].items(), key=lambda x: x[1], reverse=True):
            lines.append(f"- {category}: {amount:.2f} {CURRENCY}")
    else:
        lines.append("- Ma'lumot yo'q")

    lines.append("")
    lines.append("Kunlar bo'yicha breakdown:")
    has_daily = False
    for date_key, item in data["daily_breakdown"].items():
        if item["work_hours"] == 0 and item["uy_total"] == 0 and item["ozim_total"] == 0:
            continue
        has_daily = True
        lines.append(
            f"- {date_key}: {format_hours(item['work_hours'])} | Uy: {item['uy_total']:.2f} | O'zim: {item['ozim_total']:.2f}"
        )
    if not has_daily:
        lines.append("- Ma'lumot yo'q")

    draw_multiline_pdf_text(pdf, lines)
    pdf.save()
    return file_path


async def send_export_files(chat_id: int, context: ContextTypes.DEFAULT_TYPE, title: str, data: dict, period_name: str) -> None:
    excel_path = create_excel_report(chat_id, title, data, period_name)
    pdf_path = create_pdf_report(chat_id, title, data, period_name)

    try:
        with open(excel_path, "rb") as f:
            await context.bot.send_document(
                chat_id=chat_id,
                document=InputFile(f, filename=excel_path.name),
                caption=f"{title} - Excel",
            )
        with open(pdf_path, "rb") as f:
            await context.bot.send_document(
                chat_id=chat_id,
                document=InputFile(f, filename=pdf_path.name),
                caption=f"{title} - PDF",
            )
    finally:
        if excel_path.exists():
            excel_path.unlink(missing_ok=True)
        if pdf_path.exists():
            pdf_path.unlink(missing_ok=True)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    ensure_user(update.effective_chat.id, user.full_name, user.username)
    text = (
        "Assalomu alaykum! Men ish soati va xarajatlarni hisoblaydigan botman.\n\n"
        "Asosiy buyruqlar:\n"
        "/work - ishlagan soatni kiritish\n"
        "/expense - xarajat qo'shish\n"
        "/today - bugungi hisobot\n"
        "/week - haftalik hisobot\n"
        "/month - oylik hisobot\n"
        "/expenses - oxirgi xarajatlar\n"
        "/exportweek - haftalik Excel va PDF\n"
        "/exportmonth - oylik Excel va PDF\n"
        "/cancel - joriy amalni bekor qilish"
    )
    await update.message.reply_text(text, reply_markup=MAIN_KEYBOARD)


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "Buyruqlar:\n"
        "/work - ishlagan soatni kiritish\n"
        "/expense - xarajat qo'shish\n"
        "/today - bugungi hisobot\n"
        "/week - haftalik hisobot\n"
        "/month - oylik hisobot\n"
        "/expenses - oxirgi xarajatlar\n"
        "/exportweek - haftalik Excel va PDF\n"
        "/exportmonth - oylik Excel va PDF\n"
        "/cancel - bekor qilish",
        reply_markup=MAIN_KEYBOARD,
    )


async def work_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    chat_id = update.effective_chat.id
    user = update.effective_user
    ensure_user(chat_id, user.full_name, user.username)
    await update.message.reply_text(
        "Bugun necha soat ishlading? Masalan: 4 yoki 6.5",
        reply_markup=ReplyKeyboardRemove(),
    )
    return WORK_HOURS


async def work_hours(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.replace(",", ".").strip()
    try:
        hours = float(raw)
        if hours <= 0 or hours > 24:
            raise ValueError
    except ValueError:
        await update.message.reply_text("Iltimos, to'g'ri soat kiriting. Masalan: 4 yoki 7.5")
        return WORK_HOURS

    context.user_data["work_hours"] = hours
    await update.message.reply_text("Qayerda ishlading? Masalan: uy, ofis, kutubxona, kafe")
    return WORK_LOCATION


async def work_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    location = update.message.text.strip()
    if not location:
        await update.message.reply_text("Joy nomi bo'sh bo'lmasin.")
        return WORK_LOCATION

    hours = float(context.user_data["work_hours"])
    add_work_entry(update.effective_chat.id, hours, location, now_local())

    context.user_data.pop("work_hours", None)
    await update.message.reply_text(
        f"Ish vaqti saqlandi: {format_hours(hours)} | Joy: {location}",
        reply_markup=MAIN_KEYBOARD,
    )
    return ConversationHandler.END


async def expense_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Xarajat qaysi guruhga kiradi?", reply_markup=EXPENSE_GROUP_KEYBOARD)
    return EXPENSE_GROUP


async def expense_group(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    if raw not in {"🏠 Uy uchun", "👤 O'zim uchun"}:
        await update.message.reply_text("Iltimos, pastdagi variantlardan birini tanla.")
        return EXPENSE_GROUP

    group = "Uy uchun" if raw == "🏠 Uy uchun" else "O'zim uchun"
    context.user_data["expense_group"] = group
    await update.message.reply_text(
        f"Summa yubor. Masalan: 12.50 ({CURRENCY})",
        reply_markup=ReplyKeyboardRemove(),
    )
    return EXPENSE_AMOUNT


async def expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.replace(",", ".").strip()
    try:
        amount = float(raw)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("Iltimos, to'g'ri summa yubor. Masalan: 8.99")
        return EXPENSE_AMOUNT

    context.user_data["expense_amount"] = amount
    await update.message.reply_text(
        "Kategoriya kiriting. Masalan: oziq-ovqat, transport, ijara, dori, coffee"
    )
    return EXPENSE_CATEGORY


async def expense_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    category = update.message.text.strip()
    if not category:
        await update.message.reply_text("Kategoriya bo'sh bo'lmasin.")
        return EXPENSE_CATEGORY

    context.user_data["expense_category"] = category
    await update.message.reply_text(
        "Izoh yoki qayerga ishlatilgani haqida yoz. Yozmasang 'skip' deb yubor.\n"
        "Masalan: bozordan mahsulot, taksi, tushlik"
    )
    return EXPENSE_NOTE


async def expense_note(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    note = update.message.text.strip()
    if note.lower() == "skip":
        note = ""

    add_expense(
        chat_id=update.effective_chat.id,
        expense_group=context.user_data["expense_group"],
        amount=float(context.user_data["expense_amount"]),
        category=context.user_data["expense_category"],
        note=note,
        spent_at=now_local(),
    )

    await update.message.reply_text(
        f"Xarajat saqlandi: {context.user_data['expense_group']} | "
        f"{context.user_data['expense_amount']:.2f} {CURRENCY} | "
        f"{context.user_data['expense_category']}",
        reply_markup=MAIN_KEYBOARD,
    )

    context.user_data.pop("expense_group", None)
    context.user_data.pop("expense_amount", None)
    context.user_data.pop("expense_category", None)
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("work_hours", None)
    context.user_data.pop("expense_group", None)
    context.user_data.pop("expense_amount", None)
    context.user_data.pop("expense_category", None)
    await update.message.reply_text("Amal bekor qilindi.", reply_markup=MAIN_KEYBOARD)
    return ConversationHandler.END


async def today_summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    start_dt, end_dt = today_range(now_local())
    data = build_summary_data(update.effective_chat.id, start_dt, end_dt)
    await update.message.reply_text(format_summary("📅 Bugungi hisobot", data))


async def week_summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    start_dt, end_dt = week_range(now_local())
    data = build_summary_data(update.effective_chat.id, start_dt, end_dt)
    await update.message.reply_text(format_summary("📊 Haftalik hisobot", data))


async def month_summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    start_dt, end_dt = month_range(now_local())
    data = build_summary_data(update.effective_chat.id, start_dt, end_dt)
    await update.message.reply_text(format_summary("🗓 Oylik hisobot", data))


async def list_recent_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    rows = get_recent_expenses(update.effective_chat.id, 10)
    if not rows:
        await update.message.reply_text("Hozircha xarajatlar yo'q.")
        return

    lines = ["🧾 Oxirgi xarajatlar:"]
    for row in rows:
        note = f" | Izoh: {row['note']}" if row["note"] else ""
        lines.append(
            f"- {format_dt(parse_iso(row['spent_at']))} | {row['expense_group']} | "
            f"{row['category']} | {float(row['amount']):.2f} {CURRENCY}{note}"
        )
    await update.message.reply_text("\n".join(lines))


async def export_week(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    start_dt, end_dt = week_range(now_local())
    data = build_summary_data(chat_id, start_dt, end_dt)
    await update.message.reply_text("Haftalik Excel va PDF tayyorlanyapti...")
    await send_export_files(chat_id, context, "Haftalik hisobot", data, "week")


async def export_month(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    start_dt, end_dt = month_range(now_local())
    data = build_summary_data(chat_id, start_dt, end_dt)
    await update.message.reply_text("Oylik Excel va PDF tayyorlanyapti...")
    await send_export_files(chat_id, context, "Oylik hisobot", data, "month")


async def menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = update.message.text.strip()

    if text == "📅 Bugun":
        await today_summary(update, context)
    elif text == "📊 Hafta":
        await week_summary(update, context)
    elif text == "🗓 Oy":
        await month_summary(update, context)
    elif text == "🧾 Oxirgi xarajatlar":
        await list_recent_expenses(update, context)
    elif text == "📤 Haftalik export":
        await export_week(update, context)
    elif text == "📤 Oylik export":
        await export_month(update, context)


async def daily_reminder_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    for chat_id in get_all_user_ids():
        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text=(
                    "⏰ Eslatma: bugungi ishlagan soating va xarajatlaringni kiritishni unutmang.\n"
                    "Bugun necha soat ishlaganing va qayerda ishlaganingni yozib qo'y."
                ),
                reply_markup=MAIN_KEYBOARD,
            )
        except Exception as exc:
            logger.warning("Daily reminder yuborilmadi chat_id=%s: %s", chat_id, exc)


async def scheduled_reports_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    current = now_local()
    is_sunday = current.weekday() == 6
    is_month_end = current.day == last_day_of_month(current)

    for chat_id in get_all_user_ids():
        try:
            if is_sunday:
                week_start, week_end = week_range(current)
                week_data = build_summary_data(chat_id, week_start, week_end)
                week_title = "📊 Avtomatik haftalik hisobot"
                await context.bot.send_message(chat_id=chat_id, text=format_summary(week_title, week_data))
                await send_export_files(chat_id, context, week_title, week_data, "auto_week")

            if is_month_end:
                month_start, month_end = month_range(current)
                month_data = build_summary_data(chat_id, month_start, month_end)
                month_title = "🗓 Avtomatik oylik hisobot"
                await context.bot.send_message(chat_id=chat_id, text=format_summary(month_title, month_data))
                await send_export_files(chat_id, context, month_title, month_data, "auto_month")
        except Exception as exc:
            logger.warning("Scheduled report yuborilmadi chat_id=%s: %s", chat_id, exc)


def schedule_jobs(application: Application) -> None:
    if application.job_queue is None:
        logger.warning("JobQueue mavjud emas. Rejalashtirilgan xabarlar ishlamaydi.")
        return

    application.job_queue.run_daily(
        daily_reminder_job,
        time=time(hour=20, minute=0, tzinfo=LOCAL_TZ),
        name="daily_reminder_20_00",
    )
    application.job_queue.run_daily(
        scheduled_reports_job,
        time=time(hour=21, minute=0, tzinfo=LOCAL_TZ),
        name="scheduled_reports_21_00",
    )


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Update handling vaqtida xatolik yuz berdi:", exc_info=context.error)


def main() -> None:
    if TOKEN == "YOUR_BOT_TOKEN_HERE":
        raise ValueError("TELEGRAM_BOT_TOKEN ni environment variable sifatida o'rnating.")

    init_db()

    application = Application.builder().token(TOKEN).build()

    work_conv = ConversationHandler(
        entry_points=[
            CommandHandler("work", work_start),
            MessageHandler(filters.Regex(r"^🕒 Ish soati qo'shish$"), work_start),
        ],
        states={
            WORK_HOURS: [MessageHandler(filters.TEXT & ~filters.COMMAND, work_hours)],
            WORK_LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, work_location)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            MessageHandler(filters.Regex(r"^❌ Bekor qilish$"), cancel),
        ],
        allow_reentry=True,
    )

    expense_conv = ConversationHandler(
        entry_points=[
            CommandHandler("expense", expense_start),
            MessageHandler(filters.Regex(r"^💸 Xarajat qo'shish$"), expense_start),
        ],
        states={
            EXPENSE_GROUP: [MessageHandler(filters.TEXT & ~filters.COMMAND, expense_group)],
            EXPENSE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, expense_amount)],
            EXPENSE_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, expense_category)],
            EXPENSE_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, expense_note)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            MessageHandler(filters.Regex(r"^❌ Bekor qilish$"), cancel),
        ],
        allow_reentry=True,
    )

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("work", work_start))
    application.add_handler(CommandHandler("today", today_summary))
    application.add_handler(CommandHandler("week", week_summary))
    application.add_handler(CommandHandler("month", month_summary))
    application.add_handler(CommandHandler("expenses", list_recent_expenses))
    application.add_handler(CommandHandler("exportweek", export_week))
    application.add_handler(CommandHandler("exportmonth", export_month))
    application.add_handler(CommandHandler("cancel", cancel))

    application.add_handler(work_conv)
    application.add_handler(expense_conv)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_router), group=2)

    application.add_error_handler(error_handler)
    schedule_jobs(application)

    logger.info("Bot ishga tushdi...")
    application.run_polling()


if __name__ == "__main__":
    main()
